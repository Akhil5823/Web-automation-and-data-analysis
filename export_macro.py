import xlwings as xw
# import glob
import os.path
# import  jpype  
from zipfile import ZipFile   
import  asposecells
# jpype.startJVM() 
import win32com.client as win32
# from asposecells.api import Workbook
import conn_site
import time
import shutil
import openpyxl
import os
import warnings
from openpyxl import load_workbook

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import shutil
#copying the newly added file 

original = r'Kaarunya Files Final 10 Feb\export1.xlsx'
target = r'Kaarunya Files Final 10 Feb\export.xlsx'
shutil.copyfile(original, target)
time.sleep(5)

#applying utilisation macro

try:
    print('3')
    wb1=  xw.Book(r"Kaarunya Files Final 10 Feb\export.xlsx")
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module1.Utilisation_ModifieD_New")
    wb1.activate("Data")
    macro1()
    wb1.close()
    wb2.close()

    print("utilisation macro applied successfully !!")
except Exception as e:
    print(f"An error occurred: {e}")






# # #resource distribution  BA macro on export

try:
    print('3')
    wb1=  xw.Book(r"Kaarunya Files Final 10 Feb\export.xlsx")
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module5.SA_Resource_Distribution_Macro_UAT")
    wb1.activate("Data")
    macro1()
    # wb1.close()
    # wb2.close()

    print("resource distribution macro FOR BA applied successfully !!")
except Exception as e:
    # print("resource distribution macro FOR BA applied successfully !!")
    
    print(f"An error occurred: {e}")

time.sleep(10)

#Resource distribution SA on export

try:
    print('3')
    wb1=  xw.Book(r"Kaarunya Files Final 10 Feb\export.xlsx")
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module5.SA_Resource_Distribution_Macro_FOR_SA_REP_UAT")
    wb1.activate("Data")
    macro1()
    time.sleep(5)
    wb1.save()
    wb1.close()
    # wb2.close()

    print("resource distribution macro FOR SA applied successfully !!")
except Exception as e:
    # print("resource distribution macro FOR SA applied successfully !!")
    print(f"An error occurred: {e}")


#Retreiving the orginal state on export

time.sleep(5)
original = r'Kaarunya Files Final 10 Feb\export1.xlsx'
target = r'Kaarunya Files Final 10 Feb\export.xlsx'
shutil.copyfile(original, target)
time.sleep(5)
wb1=  xw.Book(r"Kaarunya Files Final 10 Feb\export.xlsx")
print("preping")
for sheet in wb1.sheets:
    # print(sheet)
    if sheet.name !="Data" and  sheet.name!="Pivot Sheet":
        sheet.delete()
        print("sheet deleted")
wb1.save()
wb1.close()

# # # #................................................................................................................................
# #...................................................Step 5 and 6..........................................................................
# #copying export data from first export file to required file
# print("Doing step 5 and 6.......")
# # workbook1 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\export.xlsm") #destination
# # workbook2 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\export.xlsx") #source
# # workbook1.getWorksheets().get("Data").copy(workbook2.getWorksheets().get(0))
# # workbook1.getWorksheets().get("Pivot Sheet").copy(workbook2.getWorksheets().get(1))
# # workbook1.save(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\export.xlsm")

# print("stareted copying")
# # original = r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\export.xlsx"
# # target = r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\export.xlsm"
# # shutil.copyfile(original, target)
# # time.sleep(5)
# # print("copied to dest")


# #making a copy of export file in tab to pow folder
# # import shutil
# # original = r'Kaarunya Files Final 10 Feb\export1.xlsx'
# # target = r'Kaarunya Files Final 10 Feb\export.xlsx'
# # shutil.copyfile(original, target)
# # time.sleep(5)

# #Applying time part 1 macro.............
# # wb=  xw.Book(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\export.xlsx")
# # for sheet in wb.sheets:
# #     if 'Evaluation Warning' in sheet.name:
# #         sheet.delete()
# #         # print("sheet deleted")

# # macro1=wb.macro("Module3.total_time_part1_new")
# # wb.sheets["Data"].activate()
# # macro1()
# # time.sleep(10)
# # wb.save()
# # wb.close() 
# # print("time part1 applied")


#Step 5 and 6 (time part1, time part 2, )
try:
    print('3')
    wb1=  xw.Book(r"Kaarunya Files Final 10 Feb\export.xlsx")
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module3.total_time_part1_new")
    wb1.sheets["Data"].activate()
    macro1()
    # workbook1 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\JIRA_data.xlsx")  
    # workbook1.save()
    wb1.save()
    wb1.close()
    wb2.close()

    print("time part 1 applied successfully !!")
except Exception as e:
    print(f"An error occurred: {e}")

#changing the semicolon in row44 col27
# open the Excel file
time.sleep(10)
wb4 = openpyxl.load_workbook(r"Kaarunya Files Final 10 Feb\export.xlsx")
# select the sheet to modify
sheet = wb4['Total Time']
# change the header name
sheet.cell(row=44, column=28).value = 'Adj 4 T&T Delayed Cost ICRRB'
# save the changes
wb4.save(r"Kaarunya Files Final 10 Feb\export.xlsx")
time.sleep(5)
wb4.close()



#.........copying data from AB to Ac.........


wb = xw.Book(r"Kaarunya Files Final 10 Feb\export.xlsx")
my_values = wb.sheets['Total Time'].range('AB2:AB63').options(ndim=2).value 
wb.sheets['Total Time'].range('AC1:AC62').value = my_values
time.sleep(2) 
wb.save()

print("copied to ac")
#Applying time part 2 macro................


try:
    print('3')
    wb1=  xw.Book(r'Kaarunya Files Final 10 Feb\export.xlsx')
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module3.total_time_part2")
    wb1.sheets["Total Time"].activate()
    macro1()
    # workbook1 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\JIRA_data.xlsx")  
    # workbook1.save()
    wb1.save()
    wb1.close()
    # wb2.close()

    print("Part2 macro applied successfully !!")
except Exception as e:
    print(f"An error occurred: {e}")

# 
original = r'Kaarunya Files Final 10 Feb\export.xlsx'
target = r"Kaarunya Files Final 10 Feb\after_part2.xlsx"
shutil.copyfile(original, target)

time.sleep(4)
# quit()

#Applying ytdfeb macro...................

# time.sleep(5)
xw.Book(r"Kaarunya Files Final 10 Feb\NYTD.xlsx")
try:
    print('3')
    wb1=  xw.Book(r'Kaarunya Files Final 10 Feb\export.xlsx')
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module16.ytdfeb")
    wb1.sheets["Total Time"].activate()
    macro1()
    # workbook1 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\JIRA_data.xlsx")  
    # workbook1.save()
    time.sleep(5)
    wb1.close()
    wb2.close()

    print("ytd feb macro applied successfully !!")
except Exception as e:
    print("ytd applied")
    print(f"An error occurred: {e}")

#Applying amonth macro


try:
    print('3')
    wb1=  xw.Book(r'Kaarunya Files Final 10 Feb\export.xlsx')
    print("kk")
    for sheet in wb1.sheets:
        if 'Evaluation Warning' in sheet.name:
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module17.amonthfeb")
    wb1.sheets["Total Time"].activate()
    macro1()
    # workbook1 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\JIRA_data.xlsx")  
    # workbook1.save()
    time.sleep(5)

    wb1.close()
    wb2.close()

    print("amonth macro applied successfully !!")
except Exception as e:
    print("amonth applied")
    print(f"An error occurred: {e}")


time.sleep(5)


#......................................................................................................................................
#..............................................End of step 5 and 6......................................................................


try:
    print('3')
    wb1=  xw.Book(r"Kaarunya Files Final 10 Feb\export1.xlsx")
    print("kk")
    for sheet in wb1.sheets:
        # print(sheet)
        if sheet.name !="Data" and  sheet.name!="Pivot Sheet":
            sheet.delete()
            print("sheet deleted")
    # wb1.activate("general_report")
    wb2=xw.Book(r'Kaarunya Files Final 10 Feb\UATRJ-All-In-One-FEB_1.xlsm')
    print("applying")
    macro1=wb2.macro("Module14.rd")
    wb1.activate("Data")
    macro1()
    # workbook1 = Workbook(r"C:\Users\ebinakh\OneDrive - Ericsson\Documents\Ak_ericsson\tab_to_pow\Kaarunya Files Final 10 Feb\JIRA_data.xlsx")  
    # workbook1.save()
    wb1.close()
    wb2.close()

    print("rd macro applied successfully !!")
    print("Finishing............")
    print("Done !!")
    
except Exception as e:
    print(f"An error occurred: {e}")
time.sleep(5)










